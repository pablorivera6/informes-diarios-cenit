"""
Lectura de un submission de FastField exportado a Excel.

A diferencia del parser de Ecopetrol —que buscaba hojas por nombre fijo
(`subform_1`, `subform_2`, …) y columnas por texto exacto— este identifica cada
subformulario por su **firma de columnas**. Eso lo hace inmune a dos cosas que
rompían el original:

  * el orden en que se crean las páginas en FastField (que decide la numeración
    de los `subform_N`),
  * diferencias menores de etiqueta: mayúsculas, tildes, espacios de más,
    signos de interrogación.

Ver FORMULARIO_FASTFIELD.md para las preguntas y sus destinos.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, time

import openpyxl


# ── Normalización tolerante de etiquetas ──────────────────────────────────────

def norm(texto) -> str:
    """minúsculas, sin tildes, sin signos, espacios colapsados."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("¿", "").replace("?", "").replace("¡", "").replace("!", "")
    s = re.sub(r"[^a-z0-9#]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _buscar(fila: dict, *alternativas: str):
    """Primer valor no vacío cuya etiqueta normalizada empiece por alguna alternativa."""
    for alt in alternativas:
        clave = norm(alt)
        for k, v in fila.items():
            nk = norm(k)
            if nk == clave or nk.startswith(clave):
                if v is not None and str(v).strip() != "":
                    return v
    return None


def _num(val) -> float | None:
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    txt = str(val).strip().replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", txt)
    return float(m.group(0)) if m else None


def _hora(val) -> str:
    """Devuelve 'HH:MM'. FastField puede mandar time, datetime o texto."""
    if val is None or str(val).strip() == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, float) and 0 <= val < 1:          # fracción de día
        total = round(val * 24 * 60)
        return f"{total // 60:02d}:{total % 60:02d}"
    txt = str(val).strip()
    m = re.search(r"(\d{1,2})\s*[:.]\s*(\d{2})", txt)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if "p" in txt.lower() and h < 12:
            h += 12
        if "a" in txt.lower() and h == 12:
            h = 0
        return f"{h:02d}:{mi:02d}"
    return txt


def _fecha(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    txt = str(val).split()[0].strip()
    for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


# ── Firmas de los subformularios ──────────────────────────────────────────────
# Cada entrada: (clave, columnas_obligatorias, peso). Se elige, para cada hoja,
# la firma con más columnas obligatorias presentes.

# El prefijo "item" cubre tanto "Item de pago ejecutado" como "Items cenit",
# que es como quedó nombrado el campo en el formulario real.
_FIRMAS = [
    ("jornada",     ["hora de inicio", "hora final"]),
    ("items",       ["item"]),
    ("mano_obra",   ["cargo"]),
    ("equipos",     ["tipo de equipo"]),
    ("actividades", ["describa la actividad", "actividad"]),
    ("observaciones", ["observacion"]),
]

# Columnas de metadatos que FastField agrega a todas las hojas
_METADATOS = {
    "submission id", "formid", "form name", "submitted on", "form version",
    "submitted by", "start form time stamp", "end form time stamp",
}


def _clasificar(headers: list[str]) -> str | None:
    nh = [norm(h) for h in headers if h]
    mejor, mejor_score = None, 0
    for clave, obligatorias in _FIRMAS:
        score = sum(
            1 for o in obligatorias
            if any(h == norm(o) or h.startswith(norm(o)) for h in nh)
        )
        if score > mejor_score:
            mejor, mejor_score = clave, score
    return mejor


def _filas(ws) -> list[dict]:
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = dict(zip(headers, row))
        if any(v is not None and str(v).strip() != "" for v in fila.values()):
            out.append(fila)
    return out


# ── Parser principal ──────────────────────────────────────────────────────────

def parse_submission(file_bytes) -> dict:
    """
    Devuelve un dict con lo capturado en campo, todavía SIN resolver contra el
    libro maestro (eso lo hace utils/armado.py).
    """
    wb = openpyxl.load_workbook(file_bytes, data_only=True)

    datos: dict = {
        "fecha": None,
        "frente": "",
        "profesional": "",
        "motivos_disponibilidad": "",
        "actividades": [],
        "items": [],
        "mano_obra": [],
        "equipos": [],
        "jornadas": [],
        "observaciones": [],
        "fotos": [],
        "avisos": [],
    }

    # ── Root ────────────────────────────────────────────────────────────────
    hoja_root = next((n for n in wb.sheetnames if norm(n) == "root"), None)
    if hoja_root:
        filas = _filas(wb[hoja_root])
        root = filas[0] if filas else {}
        datos["fecha"] = _fecha(_buscar(root, "Fecha del informe", "Fecha"))
        datos["frente"] = str(
            _buscar(root, "Frente o sitio de trabajo", "Frente / Sitio", "Frente",
                    "Locacion", "Localizacion", "Sitio") or ""
        ).strip()
        datos["profesional"] = str(
            _buscar(root, "Profesional lider PCC", "Profesional Lider") or ""
        ).strip()
        datos["motivos_disponibilidad"] = str(
            _buscar(root, "Motivos de disponibilidad", "Motivos disponibilidad") or ""
        ).strip()
    else:
        datos["avisos"].append("El export no trae hoja 'Root'.")

    # ── Subformularios ──────────────────────────────────────────────────────
    for nombre in wb.sheetnames:
        if norm(nombre) == "root":
            continue
        ws = wb[nombre]
        if ws.max_row < 1:
            continue
        headers = [c.value for c in ws[1]]

        if _es_picker_fotos(headers, nombre):
            # Cada picker corresponde a un slot del registro fotográfico. El
            # número va en el nombre de la hoja, y hay que leerlo: el orden de
            # wb.sheetnames es alfabético, así que 'multiphoto_picker_10'
            # aparece justo después del 1 y su foto caería en el slot 2.
            datos["fotos"].extend(_leer_fotos(_filas(ws), _slot_del_picker(nombre)))
            continue

        clave = _clasificar(headers)
        if clave is None:
            continue

        filas = _filas(ws)
        if clave == "actividades":
            datos["actividades"].extend(_leer_actividades(filas))
        elif clave == "items":
            datos["items"].extend(_leer_items(filas, datos["avisos"]))
        elif clave == "mano_obra":
            datos["mano_obra"].extend(_leer_mano_obra(filas))
        elif clave == "equipos":
            datos["equipos"].extend(_leer_equipos(filas))
        elif clave == "jornada":
            datos["jornadas"].extend(_leer_jornadas(filas))
        elif clave == "observaciones":
            datos["observaciones"].extend(_leer_observaciones(filas))

    # Los motivos también pueden venir en la página de cierre junto a observaciones
    if not datos["motivos_disponibilidad"]:
        for nombre in wb.sheetnames:
            for fila in _filas(wb[nombre]):
                v = _buscar(fila, "Motivos de disponibilidad", "Motivos disponibilidad")
                if v:
                    datos["motivos_disponibilidad"] = str(v).strip()
                    break
            if datos["motivos_disponibilidad"]:
                break

    # Un picker por slot: ordenar por destino. Si vienen de un pool común
    # (slot=None) se respeta el orden de llegada.
    if any(f.get("slot") for f in datos["fotos"]):
        datos["fotos"].sort(key=lambda f: (f.get("slot") is None, f.get("slot") or 0))

    # ── Observaciones sueltas en Root: 'Observacion #1' ... '#N' ────────────
    if not datos["observaciones"] and hoja_root:
        filas = _filas(wb[hoja_root])
        if filas:
            sueltas = []
            for k, v in filas[0].items():
                m = re.match(r"^\s*observacion\s*#?\s*(\d+)\s*$", norm(k))
                if m and v is not None and str(v).strip():
                    sueltas.append((int(m.group(1)), str(v).strip()))
            datos["observaciones"] = [t for _, t in sorted(sueltas)]

    _validar_topes(datos)
    return datos


# ── Lectores por sección ──────────────────────────────────────────────────────

def _leer_actividades(filas) -> list[str]:
    out = []
    for f in filas:
        v = _buscar(f, "Describa la actividad ejecutada", "Actividad")
        if v and str(v).strip():
            out.append(str(v).strip())
    return out


def _leer_observaciones(filas) -> list[str]:
    out = []
    for f in filas:
        v = _buscar(f, "Observacion", "Observaciones")
        if v and str(v).strip():
            out.append(str(v).strip())
    return out


ETIQUETA_ITEM = re.compile(r"^\s*0*(\d+)\s*[—\-–]+\s*(.*)")


def parse_etiqueta_item(label: str) -> tuple[int | None, str]:
    """
    '012 — PK-152+428 · Instalación de cable  [m]'  ->  (12, 'Instalación de cable')

    El número es la llave real; la descripción solo se usa como respaldo cuando
    el número no está en el catálogo. Se aceptan las dos ubicaciones del sitio:
    delante con '·' (formato de Cenit) o al final entre paréntesis.
    """
    m = ETIQUETA_ITEM.match(str(label))
    if not m:
        return None, str(label).strip()

    desc = m.group(2).strip()
    desc = re.sub(r"\s*\[[^\]]*\]\s*$", "", desc).strip()      # quita el [m] final
    # Sitio al inicio: 'PK-152+428 · Descripción'
    desc = re.sub(r"^[^·]{0,30}·\s*", "", desc).strip()
    # Sitio al final entre paréntesis, solo si parece un PK y no parte de la
    # descripción: hay 9 ítems que terminan en '(Instalaciones temporales)' y
    # similares, y esos no se deben recortar.
    desc = re.sub(r"\s*\(\s*PK[^)]*\)\s*$", "", desc, flags=re.I).strip()
    return int(m.group(1)), desc


def _leer_items(filas, avisos) -> list[dict]:
    out = []
    for f in filas:
        etiqueta = _buscar(f, "Item de pago ejecutado", "Item de pago",
                           "Items cenit", "Item")
        if not etiqueta:
            continue
        num, desc = parse_etiqueta_item(etiqueta)
        if num is None:
            avisos.append(f"No pude leer el número de ítem en: {etiqueta!r}")
            continue

        c1 = _num(_buscar(f, "Cantidad #1", "Cantidad"))
        c2 = _num(_buscar(f, "Cantidad — dimensión 2", "Cantidad dimension 2", "Cantidad #2"))
        c3 = _num(_buscar(f, "Cantidad — dimensión 3", "Cantidad dimension 3", "Cantidad #3"))
        if c1 is None:
            continue

        cantidad = c1
        for c in (c2, c3):
            if c:
                cantidad *= c

        out.append({
            "num":         num,
            "descripcion": desc,
            "etiqueta":    str(etiqueta).strip(),
            "dimensiones": [d for d in (c1, c2, c3) if d is not None],
            "cantidad":    round(cantidad, 4),
        })
    return out


def _leer_mano_obra(filas) -> list[dict]:
    out = []
    for f in filas:
        cargo = _buscar(f, "Cargo")
        if not cargo:
            continue
        out.append({
            "cargo":      str(cargo).strip(),
            "cantidad":   _num(_buscar(f, "Cantidad", "Cant")) or 1,
            "horas":      _num(_buscar(f, "Horas laboradas", "Horas")) or 0,
            "disponible": _num(_buscar(f, "Horas disponible")) or 0,
        })
    return out


def _leer_equipos(filas) -> list[dict]:
    out = []
    for f in filas:
        tipo = _buscar(f, "Tipo de equipo")
        if not tipo:
            continue
        out.append({
            "tipo":            str(tipo).strip(),
            "cantidad":        _num(_buscar(f, "Cantidad", "Cant")) or 1,
            "horas":           _num(_buscar(f, "Horas laboradas", "Horas")) or 0,
            "disponible":      _num(_buscar(f, "Horas disponible")) or 0,
            "fuera_servicio":  _num(_buscar(f, "Horas fuera de servicio")) or 0,
        })
    return out


def _leer_jornadas(filas) -> list[dict]:
    out = []
    for f in filas:
        inicio = _hora(_buscar(f, "Hora de inicio", "Hora inicio"))
        final  = _hora(_buscar(f, "Hora final", "Hora fin"))
        if not inicio and not final:
            continue
        evento = str(_buscar(f, "Hubo algun evento de suspension", "Evento") or "").strip()
        total  = _buscar(f, "Total de horas trabajadas", "Total horas")
        out.append({
            "frente":         str(_buscar(f, "Frente") or "").strip(),
            "hora_inicio":    inicio,
            "hora_final":     final,
            "total_horas":    _fmt_total(total, inicio, final),
            "hubo_evento":    evento.upper() if evento else "",
            "evento_inicio":  _hora(_buscar(f, "Hora en que inicio el evento", "Hora inicio evento")),
            "evento_fin":     _hora(_buscar(f, "Hora en que termino el evento", "Hora fin evento")),
            "evento_desc":    str(_buscar(f, "Que ocurrio", "Descripcion del evento") or "").strip(),
        })
    return out


def _fmt_total(total, inicio: str, final: str) -> str:
    """Total de horas. Si no viene, se calcula de inicio y fin."""
    n = _num(total)
    if n is not None:
        horas, minutos = int(n), round((n - int(n)) * 60)
        return f"{horas}:{minutos:02d}"
    if inicio and final:
        try:
            hi, mi = (int(x) for x in inicio.split(":"))
            hf, mf = (int(x) for x in final.split(":"))
            delta = (hf * 60 + mf) - (hi * 60 + mi)
            if delta < 0:
                delta += 24 * 60
            return f"{delta // 60}:{delta % 60:02d}"
        except ValueError:
            pass
    return str(total).strip() if total else ""


def _es_picker_fotos(headers, nombre) -> bool:
    nh = {norm(h) for h in headers if h}
    return "photo" in nh or "multiphoto" in norm(nombre) or "foto" in norm(nombre)


_NUM_PICKER = re.compile(r"(\d+)\s*$")


def _slot_del_picker(nombre: str) -> int | None:
    """'multiphoto_picker_7' -> 7. None si el nombre no lleva número."""
    m = _NUM_PICKER.search(nombre)
    return int(m.group(1)) if m else None


def _leer_fotos(filas, slot: int | None = None) -> list[dict]:
    out = []
    for f in filas:
        archivo = _buscar(f, "Photo", "Foto", "Imagen")
        if not archivo:
            continue
        out.append({
            "filename":    str(archivo).strip(),
            "descripcion": str(
                _buscar(f, "Comment", "Descripcion de la foto", "Descripcion") or ""
            ).strip(),
            # Slot destino cuando el formulario tiene un picker por posición.
            # None = pool común, se colocan en orden de llegada.
            "slot":        slot,
        })
    return out


# ── Topes del formato ─────────────────────────────────────────────────────────

TOPES = {
    "actividades":   (28, "actividades ejecutadas"),
    "jornadas":      (2,  "filas de jornada de trabajo"),
    "observaciones": (8,  "observaciones"),
    "fotos":         (10, "fotos"),
}


def _validar_topes(datos: dict):
    for clave, (tope, etiqueta) in TOPES.items():
        n = len(datos.get(clave) or [])
        if n > tope:
            datos["avisos"].append(
                f"Llegaron {n} {etiqueta} y el formato solo tiene {tope}. "
                f"Se usarán las primeras {tope}."
            )
