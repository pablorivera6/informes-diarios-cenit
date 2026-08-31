"""
Capa de mapeo del informe diario de Cenit.

Equivale a `utils/excel_ops.py` del proyecto Ecopetrol, pero para el formato
GDA-FR-363 (contrato 8000008746 / ODS03).

Estructura del libro
────────────────────
  1. Informe Diario   visible   El formato que se entrega
  2. Reg.Fotográfico  visible   10 slots de foto (2 columnas x 5 filas)
  Costo Real PDT      OCULTA    Matriz maestra de captura: ítem x día
  RESUMEN PXQ         OCULTA    Agregación PxQ — SOLO LECTURA, nunca se escribe
  EQUIPOS             OCULTA    Matriz equipo x día
  HH                  OCULTA    Matriz cargo x día

El motor del informe
────────────────────
`1. Informe Diario`!Z6 es el CONSECUTIVO y es el único input de navegación.
De ahí sale la fecha por HLOOKUP, y de la fecha salen por HLOOKUP los datos de
las tres matrices ocultas.

El consecutivo es el día calendario corrido desde DIA1 (06-abr-2026 = 1), así
que la columna de cada matriz se calcula por aritmética directa en vez de
escanear la fila de fechas.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter

from utils.zip_writer import XlsxZipWriter

# ── Nombres de hoja ───────────────────────────────────────────────────────────

SH_INFORME = "1. Informe Diario"
SH_FOTOS   = "2. Reg.Fotográfico"
SH_COSTO   = "Costo Real PDT"
SH_PXQ     = "RESUMEN PXQ"
SH_EQUIPOS = "EQUIPOS"
SH_HH      = "HH"

# ── Anclaje de fechas ─────────────────────────────────────────────────────────

# Día 1 del proyecto. Se lee de 'Costo Real PDT'!N3; esto es solo el respaldo.
DIA1_FALLBACK = date(2026, 4, 6)

# Columna (1-indexed) del día 1 en cada matriz
_COL_DIA1 = {
    SH_COSTO:   14,   # N
    SH_PXQ:     22,   # V  (solo lectura)
    SH_HH:      16,   # P
    SH_EQUIPOS:  9,   # I
}

# ── Rangos de captura en 'Costo Real PDT' ─────────────────────────────────────

COSTO_FILA_INI = 98
COSTO_FILA_FIN = 1112
COSTO_COL_CODIGO   = 1   # A  Codigo Catalogo O item
COSTO_COL_NUM      = 2   # B  Numero Asociado  (llave contra RESUMEN PXQ!B)
COSTO_COL_TAREA    = 3   # C  Nombre de tarea
COSTO_COL_ESTACION = 4   # D  Estación
COSTO_COL_UNDFUNC  = 5   # E  Und. Funcional
COSTO_COL_CANT     = 6   # F  Cantidad Contractual
COSTO_COL_UNIDAD   = 7   # G  Und. de medida
COSTO_COL_PRECIO   = 8   # H  Precio unitario
COSTO_COL_EJECUT   = 13  # M  Cant. Ejecutada (fórmula)

# ── Rangos de las matrices de recursos ────────────────────────────────────────

HH_FILA_INI, HH_FILA_FIN = 7, 52
HH_COL_CARGO = 6         # F

EQ_FILA_INI, EQ_FILA_FIN = 7, 30
EQ_COL_TIPO = 4          # D

# ── Celdas del informe ────────────────────────────────────────────────────────

CELL_CONSECUTIVO = (6, 26)      # Z6

# Actividades ejecutadas: G662:AC689 (28 líneas), frente en B662:F689
ACT_FILA_INI, ACT_FILA_FIN = 662, 689
ACT_COL_TEXTO  = 7              # G
ACT_COL_FRENTE = 2              # B

# Mano de obra / equipo empleado: filas 693-711
REC_FILA_INI, REC_FILA_FIN = 693, 711
REC_COL_MO_DISPONIBLE = 13      # M  Horas disponible (personal)
REC_COL_EQ_DISPONIBLE = 26      # Z  Horas disponible (equipo)
REC_COL_EQ_FUERA_SERV = 28      # AB Horas fuera de servicio

CELL_MOTIVOS_DISPONIBILIDAD = (713, 2)   # B713:AC714

# Jornada de trabajo: filas 718-719 (solo 2 frentes en la plantilla)
JOR_FILA_INI, JOR_FILA_FIN = 718, 719
JOR_COLS = {
    "frente":          2,    # B
    "hora_inicio":     7,    # G
    "hora_final":      9,    # I
    "total_horas":    11,    # K
    "hubo_evento":    13,    # M
    "evento_inicio":  15,    # O
    "evento_fin":     17,    # Q
    "evento_desc":    19,    # S
}

# Observaciones: B722:AC729 (8 líneas)
OBS_FILA_INI, OBS_FILA_FIN = 722, 729
OBS_COL = 2                     # B

# ── Registro fotográfico ──────────────────────────────────────────────────────

# (fila_ini, col_ini, fila_fin, col_fin, fila_desc, col_desc)  — todo 1-indexed
PHOTO_SLOTS = [
    (12,  2, 22, 16, 23,  6),   # 1  B12:P22   desc F23
    (12, 18, 22, 29, 23, 21),   # 2  R12:AC22  desc U23
    (25,  2, 35, 16, 36,  6),   # 3  B25:P35   desc F36
    (25, 18, 35, 29, 36, 21),   # 4  R25:AC35  desc U36
    (38,  2, 48, 16, 49,  6),   # 5  B38:P48   desc F49
    (38, 18, 48, 29, 49, 21),   # 6  R38:AC48  desc U49
    (51,  2, 61, 16, 62,  6),   # 7  B51:P61   desc F62
    (51, 18, 61, 29, 62, 21),   # 8  R51:AC61  desc U62
    (64,  2, 73, 16, 75,  6),   # 9  B64:P73   desc F75
    (64, 18, 74, 29, 75, 21),   # 10 R64:AC74  desc U75
]
MAX_FOTOS = len(PHOTO_SLOTS)

FOTO_NAME_PREFIX = "PCC_FOTO_"

# Cómo encajar la foto en la caja. Las cajas de la plantilla son tiras anchas
# (~555x206 px, relación 2.7:1) porque el formato nunca tuvo fotos y las filas
# quedaron a altura por defecto. Una foto 4:3 en modo "contener" deja casi la
# mitad de la caja en blanco.
#   "contener" — escala completa, centra sobre blanco. No pierde nada. (default)
#   "llenar"   — recorta los bordes para llenar la caja. Se ve mejor, pero
#                sacrifica los extremos de la foto.
MODO_FOTO = "contener"
_EMU_PER_PX = 9525
_INSET_EMU  = 19050          # ~2 px de margen para no tapar los bordes

# Las horas de la jornada se escriben como texto ("07:00"). La plantilla es
# inconsistente — fila 718 tiene formato h:mm y fila 719 General — así que un
# número se vería como "0,29" en una de las dos. Texto se ve bien en ambas.
JORNADA_HORAS_COMO_TEXTO = True


# ══ Fechas y consecutivos ═════════════════════════════════════════════════════

def leer_dia1(wb) -> date:
    """Día 1 del proyecto, leído de 'Costo Real PDT'!N3."""
    try:
        v = wb[SH_COSTO].cell(row=3, column=14).value
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
    except Exception:
        pass
    return DIA1_FALLBACK


def consecutivo_a_fecha(consecutivo: int, dia1: date = DIA1_FALLBACK) -> date:
    return dia1 + timedelta(days=consecutivo - 1)


def fecha_a_consecutivo(f, dia1: date = DIA1_FALLBACK) -> int:
    if isinstance(f, datetime):
        f = f.date()
    return (f - dia1).days + 1


def columna_del_dia(sheet: str, consecutivo: int) -> int:
    """Columna (1-indexed) que corresponde al consecutivo en la matriz dada."""
    if sheet not in _COL_DIA1:
        raise KeyError(f"'{sheet}' no es una matriz por días")
    if consecutivo < 1:
        raise ValueError(f"consecutivo inválido: {consecutivo}")
    return _COL_DIA1[sheet] + consecutivo - 1


# ══ Lectura del libro ═════════════════════════════════════════════════════════

def leer_consecutivo_actual(wb) -> int:
    """Consecutivo del informe cargado ('1. Informe Diario'!Z6)."""
    v = wb[SH_INFORME].cell(row=CELL_CONSECUTIVO[0], column=CELL_CONSECUTIVO[1]).value
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def leer_items(wb) -> list[dict]:
    """
    Catálogo de ítems contractuales desde 'Costo Real PDT'.

    Ojo con los tamaños: 'RESUMEN PXQ' tiene 597 filas y 'Costo Real PDT' 1015,
    pero son casi todas relleno pre-formateado para expansión futura. El
    catálogo real del contrato son ~92 filas (~88 ítems pagables + 4
    encabezados de alcance), y termina alrededor de la fila 188.

    Devuelve solo las filas con Numero Asociado y descripción — las filas de
    encabezado de alcance ("ALCANCE B SERVICIOS") se marcan con es_encabezado.
    """
    ws = wb[SH_COSTO]
    items: list[dict] = []

    for row in range(COSTO_FILA_INI, COSTO_FILA_FIN + 1):
        num = ws.cell(row=row, column=COSTO_COL_NUM).value
        if not isinstance(num, (int, float)) or num != int(num):
            continue
        tarea = ws.cell(row=row, column=COSTO_COL_TAREA).value
        if not tarea or str(tarea).strip() in ("", "0"):
            continue

        cant = ws.cell(row=row, column=COSTO_COL_CANT).value
        und  = ws.cell(row=row, column=COSTO_COL_UNIDAD).value
        items.append({
            "row_num":        row,
            "num":            int(num),
            "codigo":         str(ws.cell(row=row, column=COSTO_COL_CODIGO).value or ""),
            "descripcion":    str(tarea).strip(),
            "estacion":       ws.cell(row=row, column=COSTO_COL_ESTACION).value or "",
            "und_funcional":  ws.cell(row=row, column=COSTO_COL_UNDFUNC).value or "",
            "cantidad_total": cant if isinstance(cant, (int, float)) else 0,
            "unidad":         str(und).strip() if und else "",
            "precio_unit":    ws.cell(row=row, column=COSTO_COL_PRECIO).value or 0,
            "ejecutado":      ws.cell(row=row, column=COSTO_COL_EJECUT).value or 0,
            "es_encabezado":  not (und and str(und).strip() not in ("", "0")),
        })

    return items


def leer_cargos(wb) -> list[dict]:
    """Filas de la matriz HH con cargo definido."""
    ws = wb[SH_HH]
    out = []
    for row in range(HH_FILA_INI, HH_FILA_FIN + 1):
        cargo = ws.cell(row=row, column=HH_COL_CARGO).value
        if cargo and str(cargo).strip():
            out.append({
                "row_num": row,
                "cargo":   str(cargo).strip(),
                "nombre":  ws.cell(row=row, column=5).value or "",
            })
    return out


def leer_equipos(wb) -> list[dict]:
    """Filas de la matriz EQUIPOS con tipo definido."""
    ws = wb[SH_EQUIPOS]
    out = []
    for row in range(EQ_FILA_INI, EQ_FILA_FIN + 1):
        tipo = ws.cell(row=row, column=EQ_COL_TIPO).value
        if tipo and str(tipo).strip():
            out.append({"row_num": row, "tipo": str(tipo).strip()})
    return out


def mapa_filas_recursos(wb) -> tuple[dict[int, int], dict[int, int]]:
    """
    Correspondencia entre la fila de cada matriz oculta y su fila en el bloque
    de recursos del informe (693-711).

    El bloque del informe repite las etiquetas de HH (columna B) y de EQUIPOS
    (columna O) en el mismo orden, pero se hace el cruce por TEXTO y no por
    posición: si alguien inserta una fila en una de las matrices, el cruce
    posicional escribiría las horas disponible en el cargo equivocado.

    Devuelve ({fila_HH: fila_informe}, {fila_EQUIPOS: fila_informe}).
    """
    ws = wb[SH_INFORME]

    def _cruce(catalogo, campo, col_informe):
        etiquetas = {}
        for fila in range(REC_FILA_INI, REC_FILA_FIN + 1):
            v = ws.cell(row=fila, column=col_informe).value
            if v and str(v).strip():
                etiquetas.setdefault(str(v).strip().lower(), []).append(fila)

        usados, mapa = set(), {}
        for i, c in enumerate(catalogo):
            clave = str(c[campo]).strip().lower()
            destino = next((f for f in etiquetas.get(clave, []) if f not in usados), None)
            if destino is None:
                # Sin coincidencia de texto: caer al orden posicional
                destino = REC_FILA_INI + i
            if REC_FILA_INI <= destino <= REC_FILA_FIN:
                mapa[c["row_num"]] = destino
                usados.add(destino)
        return mapa

    return (
        _cruce(leer_cargos(wb), "cargo", 2),     # columna B
        _cruce(leer_equipos(wb), "tipo", 15),    # columna O
    )


# ══ Geometría de fotos ════════════════════════════════════════════════════════

def _col_px(ws, col: int) -> float:
    """Ancho de columna en píxeles (aprox. Calibri 11)."""
    letra = get_column_letter(col)
    dim = ws.column_dimensions.get(letra)
    ancho = None
    if dim is not None and dim.width:
        ancho = dim.width
    else:
        # openpyxl no expande rangos min..max, hay que buscarlos
        for d in ws.column_dimensions.values():
            if d.min <= col <= d.max and d.width:
                ancho = d.width
                break
    if ancho is None:
        ancho = ws.sheet_format.defaultColWidth or 8.43
    return round(ancho * 7) + 5


def _row_px(ws, row: int) -> float:
    """Alto de fila en píxeles."""
    dim = ws.row_dimensions.get(row)
    alto = dim.height if dim is not None and dim.height else (
        ws.sheet_format.defaultRowHeight or 15.0
    )
    return alto * 4.0 / 3.0


def medir_slot(wb, slot_idx: int) -> tuple[int, int]:
    """Tamaño en píxeles de la caja de foto (ancho, alto)."""
    ws = wb[SH_FOTOS]
    r0, c0, r1, c1, _, _ = PHOTO_SLOTS[slot_idx]
    ancho = sum(_col_px(ws, c) for c in range(c0, c1 + 1))
    alto  = sum(_row_px(ws, r) for r in range(r0, r1 + 1))
    return int(round(ancho)), int(round(alto))


def ajustar_imagen(image_bytes: bytes, box_px: tuple[int, int],
                   modo: str | None = None) -> bytes:
    """
    Encaja la foto en la caja sin deformarla. Devuelve PNG del tamaño exacto
    de la caja, para que el anclaje pueda estirarse sin distorsión.

    modo="contener" (default) escala completa y centra sobre blanco;
    modo="llenar" recorta los bordes para cubrir toda la caja.
    """
    from PIL import Image, ImageOps

    img = Image.open(io.BytesIO(image_bytes))
    img = ImageOps.exif_transpose(img)      # respeta orientación de la cámara
    img = img.convert("RGB")

    ancho, alto = max(1, box_px[0]), max(1, box_px[1])

    if (modo or MODO_FOTO) == "llenar":
        lienzo = ImageOps.fit(img, (ancho, alto), Image.LANCZOS, centering=(0.5, 0.5))
    else:
        ajustada = ImageOps.contain(img, (ancho, alto), Image.LANCZOS)
        lienzo = Image.new("RGB", (ancho, alto), (255, 255, 255))
        lienzo.paste(
            ajustada,
            ((ancho - ajustada.width) // 2, (alto - ajustada.height) // 2),
        )

    buf = io.BytesIO()
    lienzo.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


# ══ Contexto ligero ═══════════════════════════════════════════════════════════

def construir_contexto(xlsx_bytes) -> dict:
    """
    Extrae de una sola pasada todo lo que la interfaz necesita del libro, para
    NO tener que conservarlo en memoria.

    Motivo: openpyxl expande este libro a ~1 GB de RSS (las hojas ocultas son de
    848x786 y 1114x778). Guardarlo en session_state para consultar ~92 ítems,
    10 cargos y 11 equipos hacía que la app superara el límite de memoria de
    Streamlit Community Cloud. El contexto pesa unos pocos KB.
    """
    if hasattr(xlsx_bytes, "read"):
        xlsx_bytes = xlsx_bytes.read()
    wb = abrir_libro(xlsx_bytes)
    try:
        filas_mo, filas_eq = mapa_filas_recursos(wb)
        return {
            "dia1": leer_dia1(wb),
            "consecutivo_actual": leer_consecutivo_actual(wb),
            "items": [i for i in leer_items(wb) if not i["es_encabezado"]],
            "cargos": leer_cargos(wb),
            "equipos": leer_equipos(wb),
            "filas_mo": filas_mo,
            "filas_eq": filas_eq,
            "cajas": [medir_slot(wb, i) for i in range(MAX_FOTOS)],
            "ultimas_celdas": [medir_ultima_celda(wb, i) for i in range(MAX_FOTOS)],
            "dias_con_datos": _dias_con_datos(wb),
            "avisos_manuales": detectar_datos_manuales(wb),
        }
    finally:
        wb.close()


def _dias_con_datos(wb) -> dict[int, dict[str, int]]:
    """
    Cuántos registros tiene ya cada día en cada matriz.

    Se recorre con iter_rows, no celda por celda: con ws.cell() el barrido
    tardaba 21 s, así baja a medio segundo.
    """
    conteo: dict[int, dict[str, int]] = {}
    for hoja, clave, f0, f1 in (
        (SH_COSTO,   "avances", COSTO_FILA_INI, COSTO_FILA_FIN),
        (SH_HH,      "hh",      HH_FILA_INI,    HH_FILA_FIN),
        (SH_EQUIPOS, "equipos", EQ_FILA_INI,    EQ_FILA_FIN),
    ):
        ws = wb[hoja]
        base = _COL_DIA1[hoja]
        for fila in ws.iter_rows(min_row=f0, max_row=min(f1, ws.max_row),
                                 min_col=base, values_only=True):
            for i, v in enumerate(fila):
                if isinstance(v, (int, float)) and v:
                    conteo.setdefault(i + 1, {"avances": 0, "hh": 0, "equipos": 0})
                    conteo[i + 1][clave] += 1
    return conteo


def dia_ya_tiene_datos_ctx(contexto: dict, consecutivo: int) -> dict:
    return contexto["dias_con_datos"].get(
        consecutivo, {"avances": 0, "hh": 0, "equipos": 0}
    )


# ══ Limpieza del día anterior ═════════════════════════════════════════════════

# Filas 710-711 traen datos escritos a mano (Cadenero / Operador de cama Baja
# con cantidades y horas fijas) mientras el resto del bloque es fórmula. No se
# limpian automáticamente porque podrían ser un registro legítimo, pero se
# reportan para que el usuario decida.
MO_FILAS_MANUALES = (710, 711)


def detectar_datos_manuales(wb) -> list[str]:
    """
    Avisos sobre datos escritos a mano en el bloque de recursos, que la
    plantilla arrastra de un informe a otro sin que ninguna fórmula los limpie.
    """
    ws = wb[SH_INFORME]
    avisos = []
    for fila in MO_FILAS_MANUALES:
        cargo = ws.cell(row=fila, column=2).value
        cant  = ws.cell(row=fila, column=9).value    # I
        horas = ws.cell(row=fila, column=11).value   # K
        if cargo and (cant or horas):
            avisos.append(
                f"Fila {fila}: '{cargo}' tiene cantidad {cant} y {horas} h escritas "
                f"a mano (el resto del bloque es fórmula). Verifica si corresponde a hoy."
            )
    for fila in MO_FILAS_MANUALES:
        equipo = ws.cell(row=fila, column=15).value  # O
        if equipo:
            avisos.append(
                f"Fila {fila}: equipo '{equipo}' escrito a mano en la columna de tipo."
            )
    return avisos


def dia_ya_tiene_datos(wb, consecutivo: int) -> dict:
    """
    ¿La columna de ese consecutivo ya tiene avances/HH/equipos cargados?

    Los avances se ESCRIBEN SUMANDO (add_to_number), porque un mismo día puede
    recibir varios envíos de FastField para el mismo ítem. El efecto colateral
    es que regenerar dos veces el mismo consecutivo duplica las cantidades, así
    que la app debe avisar antes.
    """
    res = {"avances": 0, "hh": 0, "equipos": 0}
    try:
        col = columna_del_dia(SH_COSTO, consecutivo)
        ws = wb[SH_COSTO]
        res["avances"] = sum(
            1 for r in range(COSTO_FILA_INI, COSTO_FILA_FIN + 1)
            if isinstance(ws.cell(row=r, column=col).value, (int, float))
            and ws.cell(row=r, column=col).value
        )
        col = columna_del_dia(SH_HH, consecutivo)
        ws = wb[SH_HH]
        res["hh"] = sum(
            1 for r in range(HH_FILA_INI, HH_FILA_FIN + 1)
            if isinstance(ws.cell(row=r, column=col).value, (int, float))
            and ws.cell(row=r, column=col).value
        )
        col = columna_del_dia(SH_EQUIPOS, consecutivo)
        ws = wb[SH_EQUIPOS]
        res["equipos"] = sum(
            1 for r in range(EQ_FILA_INI, EQ_FILA_FIN + 1)
            if isinstance(ws.cell(row=r, column=col).value, (int, float))
            and ws.cell(row=r, column=col).value
        )
    except Exception:
        pass
    return res


def limpiar_dia_anterior(w: XlsxZipWriter):
    """
    Borra los bloques narrativos del informe anterior conservando estilos.

    Imprescindible: la plantilla es siempre el informe de ayer. Sin esto, las
    actividades y observaciones de ayer aparecen en el informe de hoy — las
    celdas con fórmula (que sí se refrescan solas con el consecutivo) no se
    tocan nunca.
    """
    # Actividades ejecutadas + frente
    w.clear_cell(SH_INFORME, ACT_FILA_INI, ACT_COL_FRENTE)
    for fila in range(ACT_FILA_INI, ACT_FILA_FIN + 1):
        w.clear_cell(SH_INFORME, fila, ACT_COL_TEXTO)

    # Jornada de trabajo
    for fila in range(JOR_FILA_INI, JOR_FILA_FIN + 1):
        for col in JOR_COLS.values():
            w.clear_cell(SH_INFORME, fila, col)

    # Motivos de disponibilidad
    w.clear_cell(SH_INFORME, *CELL_MOTIVOS_DISPONIBILIDAD)

    # Observaciones
    for fila in range(OBS_FILA_INI, OBS_FILA_FIN + 1):
        w.clear_cell(SH_INFORME, fila, OBS_COL)

    # Horas disponible / fuera de servicio (columnas manuales del bloque)
    for fila in range(REC_FILA_INI, REC_FILA_FIN + 1):
        for col in (REC_COL_MO_DISPONIBLE, REC_COL_EQ_DISPONIBLE, REC_COL_EQ_FUERA_SERV):
            w.clear_cell(SH_INFORME, fila, col)

    # Descripciones del registro fotográfico
    for _, _, _, _, fila_desc, col_desc in PHOTO_SLOTS:
        w.clear_cell(SH_FOTOS, fila_desc, col_desc)


# ══ Escritura del informe ═════════════════════════════════════════════════════

def generar_informe(template_bytes, datos: dict) -> bytes:
    """
    Escribe un informe diario nuevo sobre el libro maestro.

    `template_bytes`: bytes (o file-like) del último informe enviado, que hace
    de libro maestro acumulativo — igual que en el proyecto Ecopetrol.

    Claves reconocidas en `datos`:
      consecutivo        int    obligatorio
      avances            list   [{row_num, cantidad}]  -> Costo Real PDT
      hh                 list   [{row_num, horas}]     -> HH
      equipos            list   [{row_num, horas}]     -> EQUIPOS
      frente             str    B662
      actividades        list[str]                     -> G662:G689 (28 máx.)
      jornadas           list[dict]                    -> filas 718-719 (2 máx.)
      motivos_disponibilidad  str                      -> B713
      observaciones      list[str]                     -> B722:B729 (8 máx.)
      mo_disponible      list   [{row_num, horas}]     -> M693:M711
      eq_disponible      list   [{row_num, horas}]     -> Z693:Z711
      eq_fuera_servicio  list   [{row_num, horas}]     -> AB693:AB711
      fotos              list   [{image_bytes, descripcion}] (10 máx.)
      cajas              list   tamaños de caja, de construir_contexto (opcional)
      ultimas_celdas     list   px de la última celda de cada caja (opcional)
      modo_foto          str    "contener" (default) o "llenar"
    """
    if hasattr(template_bytes, "read"):
        template_bytes = template_bytes.read()

    consecutivo = int(datos["consecutivo"])
    w = XlsxZipWriter(template_bytes)

    # Sin esto el archivo se abre mostrando los datos del día anterior.
    w.force_full_recalc()

    # La plantilla es el informe de ayer: hay que borrar sus bloques narrativos
    # antes de escribir los de hoy.
    limpiar_dia_anterior(w)

    # ── Consecutivo: el único input que mueve todo el informe ────────────────
    w.set_number(SH_INFORME, CELL_CONSECUTIVO[0], CELL_CONSECUTIVO[1], consecutivo)

    # ── Matrices ocultas ─────────────────────────────────────────────────────
    col_costo = columna_del_dia(SH_COSTO, consecutivo)
    for av in datos.get("avances") or []:
        cant = av.get("cantidad") or 0
        if cant:
            w.add_to_number(SH_COSTO, int(av["row_num"]), col_costo, float(cant))

    col_hh = columna_del_dia(SH_HH, consecutivo)
    for reg in datos.get("hh") or []:
        horas = reg.get("horas") or 0
        if horas:
            w.set_number(SH_HH, int(reg["row_num"]), col_hh, float(horas))

    col_eq = columna_del_dia(SH_EQUIPOS, consecutivo)
    for reg in datos.get("equipos") or []:
        horas = reg.get("horas") or 0
        if horas:
            w.set_number(SH_EQUIPOS, int(reg["row_num"]), col_eq, float(horas))

    # ── Actividades ejecutadas ───────────────────────────────────────────────
    if datos.get("frente"):
        w.set_text(SH_INFORME, ACT_FILA_INI, ACT_COL_FRENTE, str(datos["frente"]))

    actividades = [a for a in (datos.get("actividades") or []) if str(a).strip()]
    for i, texto in enumerate(actividades[: ACT_FILA_FIN - ACT_FILA_INI + 1]):
        w.set_text(SH_INFORME, ACT_FILA_INI + i, ACT_COL_TEXTO, str(texto).strip())

    # ── Jornada de trabajo ───────────────────────────────────────────────────
    for i, j in enumerate((datos.get("jornadas") or [])[: JOR_FILA_FIN - JOR_FILA_INI + 1]):
        fila = JOR_FILA_INI + i
        for clave, col in JOR_COLS.items():
            val = j.get(clave)
            if val is None or str(val).strip() == "":
                continue
            if clave == "total_horas" and not JORNADA_HORAS_COMO_TEXTO:
                w.set_number(SH_INFORME, fila, col, float(val))
            else:
                w.set_text(SH_INFORME, fila, col, str(val).strip())

    if datos.get("motivos_disponibilidad"):
        w.set_text(SH_INFORME, *CELL_MOTIVOS_DISPONIBILIDAD,
                   str(datos["motivos_disponibilidad"]).strip())

    # ── Horas disponible / fuera de servicio ─────────────────────────────────
    for clave, col in (
        ("mo_disponible",     REC_COL_MO_DISPONIBLE),
        ("eq_disponible",     REC_COL_EQ_DISPONIBLE),
        ("eq_fuera_servicio", REC_COL_EQ_FUERA_SERV),
    ):
        for reg in datos.get(clave) or []:
            fila = int(reg["row_num"])
            if REC_FILA_INI <= fila <= REC_FILA_FIN:
                w.set_number(SH_INFORME, fila, col, float(reg.get("horas") or 0))

    # ── Observaciones ────────────────────────────────────────────────────────
    obs = [o for o in (datos.get("observaciones") or []) if str(o).strip()]
    for i, texto in enumerate(obs[: OBS_FILA_FIN - OBS_FILA_INI + 1]):
        w.set_text(SH_INFORME, OBS_FILA_INI + i, OBS_COL, str(texto).strip())

    # ── Registro fotográfico ─────────────────────────────────────────────────
    _escribir_fotos(w, datos, template_bytes)

    return w.save()


def _escribir_fotos(w: XlsxZipWriter, datos: dict, template_bytes: bytes):
    fotos = (datos.get("fotos") or [])[:MAX_FOTOS]
    if not fotos:
        return

    # Idempotencia: si se regenera el informe, no duplicar anclajes
    w.remove_pictures_named(SH_FOTOS, FOTO_NAME_PREFIX)

    # Si el llamador ya midió las cajas (construir_contexto lo hace), evitamos
    # reabrir el libro aquí: son ~500 MB de más en el momento de generar.
    cajas = datos.get("cajas")
    wb = None if cajas else openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=True)

    for idx, foto in enumerate(fotos):
        r0, c0, r1, c1, fila_desc, col_desc = PHOTO_SLOTS[idx]

        desc = str(foto.get("descripcion") or "").strip()
        if desc:
            w.set_text(SH_FOTOS, fila_desc, col_desc, desc)

        img = foto.get("image_bytes")
        if not img:
            continue

        caja = cajas[idx] if cajas else medir_slot(wb, idx)
        png = ajustar_imagen(img, caja, datos.get("modo_foto"))
        ult_col_px, ult_fila_px = (
            datos["ultimas_celdas"][idx] if datos.get("ultimas_celdas")
            else medir_ultima_celda(wb, idx)
        )

        w.insert_picture(
            SH_FOTOS,
            png,
            from_col=c0 - 1,
            from_row=r0 - 1,
            to_col=c1 - 1,
            to_row=r1 - 1,
            from_off=(_INSET_EMU, _INSET_EMU),
            to_off=(max(0, int(ult_col_px * _EMU_PER_PX) - _INSET_EMU),
                    max(0, int(ult_fila_px * _EMU_PER_PX) - _INSET_EMU)),
            name=f"{FOTO_NAME_PREFIX}{idx + 1}",
        )


def medir_ultima_celda(wb, slot_idx: int) -> tuple[float, float]:
    """Píxeles de la última columna y última fila del slot."""
    ws = wb[SH_FOTOS]
    _, _, r1, c1, _, _ = PHOTO_SLOTS[slot_idx]
    return _col_px(ws, c1), _row_px(ws, r1)


def abrir_libro(xlsx_bytes) -> openpyxl.Workbook:
    """Carga el libro con valores cacheados (para leer catálogos y fechas)."""
    if hasattr(xlsx_bytes, "read"):
        xlsx_bytes = xlsx_bytes.read()
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
