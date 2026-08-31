"""
Prueba de integración: FastField -> parser -> armado -> Excel.

Fabrica un export sintético de FastField deliberadamente hostil:
  * los subformularios van en orden barajado y con numeración que no coincide
    con el orden de las páginas del formulario,
  * las etiquetas traen mayúsculas, tildes y espacios distintos a la spec,
  * las horas llegan en tres formatos (time, fracción de día, texto "7:30 AM"),
  * hay un ítem repetido dos veces en el día (debe sumarse),
  * llegan 30 actividades y 3 jornadas (por encima de los topes de 28 y 2).

    python3 scripts/test_pipeline.py
"""
import io
import sys
from datetime import date, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from PIL import Image, ImageDraw

from utils import armado, cenit_report as cr
from utils.fastfield import parse_submission

PLANTILLA = Path.home() / "Downloads" / "2026-08-22_8000008746_ODS03_Informe_Diario_139.xlsx"
SALIDA = Path(__file__).resolve().parent.parent / "salida_prueba_pipeline.xlsx"

fallos: list[str] = []


def check(nombre, obtenido, esperado):
    ok = obtenido == esperado
    print(f"  {'PASA' if ok else 'FALLA'}  {nombre}: {obtenido!r}"
          + ("" if ok else f"  (esperado {esperado!r})"))
    if not ok:
        fallos.append(nombre)


def check_true(nombre, cond, detalle=""):
    print(f"  {'PASA' if cond else 'FALLA'}  {nombre}" + (f": {detalle}" if detalle else ""))
    if not cond:
        fallos.append(nombre)


def foto(txt):
    img = Image.new("RGB", (1200, 1600), (60, 80, 60))
    ImageDraw.Draw(img).text((60, 60), txt, fill=(255, 255, 255))
    b = io.BytesIO()
    img.save(b, format="JPEG")
    return b.getvalue()


# ══ 1. Export sintético de FastField ══════════════════════════════════════════

def construir_export() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def hoja(nombre, headers, filas):
        ws = wb.create_sheet(nombre)
        ws.append(headers)
        for f in filas:
            ws.append(f)

    # Orden barajado a propósito: la numeración de subform NO sigue las páginas
    hoja("Root",
         ["Fecha del informe", "FRENTE O SITIO DE TRABAJO ", "Motivos de disponibilidad"],
         [[date(2026, 8, 23), "PK-152+428",
           "Retroexcavadora en mantenimiento preventivo."]])

    # subform_1 = equipos (página 6)
    hoja("subform_1",
         ["Tipo de equipo", "Horas laboradas", "Horas disponible (en sitio sin operar)",
          "Horas fuera de servicio"],
         [["Camioneta 4x4", 10, 0, 0],
          ["Retroexcavadora", 0, 0, 8],
          ["Herramienta menor", 8, 0, 0]])

    # subform_2 = observaciones (página 8)
    hoja("subform_2", ["Observación"],
         [["Se presentó lluvia entre las 14:00 y 15:00."],
          ["Interventoría acompañó la jornada."]])

    # subform_3 = ítems (página 4) — etiquetas SIN tilde en "Item"
    hoja("subform_3",
         ["Item de pago ejecutado", "Cantidad", "Cantidad - dimension 2",
          "Cantidad - dimension 3"],
         # Etiquetas copiadas tal cual de catalogos/items_cenit_fastfield.csv
         [["010 — PK-152+428 · Instalación y tendido de cable de Protección Catódica  [m]",
           12.5, None, None],
          # mismo ítem otra vez: debe sumarse -> 12.5 + 7.5 = 20
          ["010 — PK-152+428 · Instalación y tendido de cable de Protección Catódica  [m]",
           7.5, None, None],
          # tres dimensiones: 2 x 1.5 x 0.8 = 2.4
          ["009 — PK-152+428 · Identificación de estructuras enterradas  [m²]",
           2, 1.5, 0.8],
          # descripción con paréntesis propios: no se deben confundir con el sitio
          ["008 — PK-152+428 · Campamento de obra (Instalaciones temporales)  [Und]",
           1, None, None],
          ["999 — Ítem que no existe en el contrato  [Und]", 1, None, None]])

    # subform_4 = jornada (página 2) — tres formatos de hora, y 3 filas (tope 2)
    hoja("subform_4",
         ["Frente", "Hora de inicio", "Hora final", "Total de horas trabajadas",
          "¿Hubo algún evento de suspensión?", "Hora en que inició el evento",
          "Hora en que terminó el evento", "¿Qué ocurrió?"],
         [["PK-152+428", time(7, 0), time(17, 0), 9, "SI", time(14, 0), time(15, 0),
           "Suspensión por tormenta eléctrica."],
          ["PK-47+128", 0.3125, "4:30 PM", None, "NO", None, None, None],
          ["PK-13+300", time(8, 0), time(12, 0), 4, "NO", None, None, None]])

    # subform_5 = mano de obra (página 5)
    hoja("subform_5",
         ["Cargo", "Horas laboradas", "Horas disponible (presente pero sin laborar)"],
         [["Director de Obra", 4, 0],
          ["Ingeniero", 9, 0],
          ["Profesional QA QC (fila 15)", 8, 0],
          ["Ayudante", 0, 6]])

    # subform_6 = actividades (página 3) — 30 filas, tope 28
    hoja("subform_6", ["Describa la actividad ejecutada"],
         [[f"Actividad número {i} ejecutada en el frente."] for i in range(1, 31)])

    hoja("multiphoto_picker_1", ["Photo", "Comment"],
         [["IMG_001.jpg", "Tendido de cable, vista general."],
          ["IMG_002.jpg", "Detalle de conexión termosoldada."],
          ["IMG_003.jpg", ""]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════

if not PLANTILLA.exists():
    sys.exit(f"No se encontró la plantilla en {PLANTILLA}")

print("\n── Parser de FastField ──────────────────────────────────────────────")
export = construir_export()
cap = parse_submission(io.BytesIO(export))

check("fecha del informe", cap["fecha"].isoformat(), "2026-08-23")
check("frente (etiqueta con mayúsculas y espacio final)", cap["frente"], "PK-152+428")
check_true("motivos de disponibilidad", cap["motivos_disponibilidad"].startswith("Retro"))
check("actividades capturadas", len(cap["actividades"]), 30)
check("ítems capturados", len(cap["items"]), 5)
check("mano de obra capturada", len(cap["mano_obra"]), 4)
check("equipos capturados", len(cap["equipos"]), 3)
check("jornadas capturadas", len(cap["jornadas"]), 3)
check("observaciones capturadas", len(cap["observaciones"]), 2)
check("fotos capturadas", len(cap["fotos"]), 3)

print("\n  Clasificación de subformularios por firma de columnas:")
print(f"    subform_1 -> equipos       ({[e['tipo'] for e in cap['equipos']]})")
print(f"    subform_3 -> ítems         (nums {[i['num'] for i in cap['items']]})")
print(f"    subform_4 -> jornada       ({[j['frente'] for j in cap['jornadas']]})")
print(f"    subform_5 -> mano de obra  ({[m['cargo'] for m in cap['mano_obra']]})")

print("\n── Normalización de horas ───────────────────────────────────────────")
j0, j1 = cap["jornadas"][0], cap["jornadas"][1]
check("hora time(7,0)", j0["hora_inicio"], "07:00")
check("hora fracción de día 0.3125", j1["hora_inicio"], "07:30")
check("hora texto '4:30 PM'", j1["hora_final"], "16:30")
check("total explícito 9 -> '9:00'", j0["total_horas"], "9:00")
check("total calculado de 07:30-16:30", j1["total_horas"], "9:00")

print("\n── Multiplicación de dimensiones ────────────────────────────────────")
it9 = next(i for i in cap["items"] if i["num"] == 9)
check("2 x 1.5 x 0.8", it9["cantidad"], 2.4)

print("\n── Etiquetas del lookup de FastField ────────────────────────────────")
it8 = next(i for i in cap["items"] if i["num"] == 8)
check("sitio al inicio se descarta de la descripción",
      it8["descripcion"], "Campamento de obra (Instalaciones temporales)")
it10 = next(i for i in cap["items"] if i["num"] == 10)
check("descripción sin el sitio ni la unidad",
      it10["descripcion"], "Instalación y tendido de cable de Protección Catódica")

print("\n── Avisos por topes excedidos ───────────────────────────────────────")
check_true("avisa 30 actividades > 28",
           any("28" in a and "actividades" in a for a in cap["avisos"]))
check_true("avisa 3 jornadas > 2",
           any("jornada" in a for a in cap["avisos"]))
for a in cap["avisos"]:
    print(f"     · {a}")

print("\n── Armado contra el libro maestro ───────────────────────────────────")
plantilla_bytes = PLANTILLA.read_bytes()
ctx = cr.construir_contexto(plantilla_bytes)
wb0 = cr.abrir_libro(plantilla_bytes)          # para las comprobaciones directas
consecutivo = cr.fecha_a_consecutivo(cap["fecha"], ctx["dia1"])
check("consecutivo derivado de la fecha", consecutivo, 140)

fotos_bytes = {"IMG_001.jpg": foto("1"), "IMG_002.jpg": foto("2"), "IMG_003.jpg": foto("3")}
datos, avisos = armado.construir(cap, ctx, consecutivo, fotos_bytes)

check("actividades recortadas al tope", len(datos["actividades"]), 28)
check("jornadas recortadas al tope", len(datos["jornadas"]), 2)
check("ítem inexistente descartado", len(datos["avances"]), 3)
check_true("avisa del ítem inexistente",
           any("999" in a or "no está en el catálogo" in a for a in avisos))
check_true("resuelve 'Profesional QA QC (fila 15)'",
           any(m["row_num"] == 15 for m in datos["_detalle"]["mano_obra"]))
check("MO con horas 0 excluida de HH", len(datos["hh"]), 3)
check("horas disponible del ayudante", datos["mo_disponible"],
      [{"row_num": 702, "horas": 6}])
check("equipo fuera de servicio (retroexcavadora)", datos["eq_fuera_servicio"],
      [{"row_num": 696, "horas": 8}])

item10 = next(i for i in datos["_detalle"]["items"] if i["num"] == 10)
avance10 = next(a for a in datos["avances"] if a["row_num"] == item10["row_num"])
check("ítem repetido sumado (12.5 + 7.5)", avance10["cantidad"], 20.0)

print("\n  Avisos del armado:")
for a in avisos:
    print(f"     · {a}")

print("\n── Generación del Excel ─────────────────────────────────────────────")
salida = cr.generar_informe(plantilla_bytes, datos)
SALIDA.write_bytes(salida)
print(f"  Generado: {SALIDA.name} ({len(salida)/1024/1024:.2f} MB)")

wb = openpyxl.load_workbook(io.BytesIO(salida), data_only=True)
inf, cos, hh, eq, fot = (wb[cr.SH_INFORME], wb[cr.SH_COSTO], wb[cr.SH_HH],
                         wb[cr.SH_EQUIPOS], wb[cr.SH_FOTOS])
col_costo = cr.columna_del_dia(cr.SH_COSTO, consecutivo)
col_hh = cr.columna_del_dia(cr.SH_HH, consecutivo)
col_eq = cr.columna_del_dia(cr.SH_EQUIPOS, consecutivo)

check("Z6", inf.cell(row=6, column=26).value, 140)
check("frente B662", inf.cell(row=662, column=2).value, "PK-152+428")
check("actividad 1", inf.cell(row=662, column=7).value,
      "Actividad número 1 ejecutada en el frente.")
check("actividad 28 (última que cabe)", inf.cell(row=689, column=7).value,
      "Actividad número 28 ejecutada en el frente.")
check("avance ítem 10", cos.cell(row=item10["row_num"], column=col_costo).value, 20.0)
check("HH ingeniero (fila 11)", hh.cell(row=11, column=col_hh).value, 9)
check("HH QA QC fila 15", hh.cell(row=15, column=col_hh).value, 8)
check("horas camioneta", eq.cell(row=7, column=col_eq).value, 10)
check("jornada 1 inicio", inf.cell(row=718, column=7).value, "07:00")
check("jornada 2 inicio (fracción)", inf.cell(row=719, column=7).value, "07:30")
check("jornada 3 descartada", inf.cell(row=720, column=2).value, None)
check("observación 1", inf.cell(row=722, column=2).value,
      "Se presentó lluvia entre las 14:00 y 15:00.")
check("desc foto 1", fot.cell(row=23, column=6).value, "Tendido de cable, vista general.")
check("foto 3 sin descripción", fot.cell(row=36, column=6).value, None)

ws_img = openpyxl.load_workbook(io.BytesIO(salida))[cr.SH_FOTOS]
check("imágenes insertadas (3 + logo)", len(ws_img._images), 4)

# ══ Simulación del camino del botón "Generar" en app.py ══════════════════════
# Es la parte que el navegador no alcanza: queda detrás del st.stop() que exige
# los dos archivos cargados.

print("\n── Camino del botón Generar (app.py) ────────────────────────────────")

# La app deja editar los textos antes de generar
actividades_editadas = "\n".join(datos["actividades"][:3] + ["Línea agregada a mano."])
observaciones_editadas = "\n".join(datos["observaciones"])
fotos_final = [
    {"image_bytes": fotos_bytes.get(f["filename"]), "descripcion": f["descripcion"]}
    for f in cap["fotos"][:10]
]

datos_final = {
    **{k: v for k, v in datos.items() if k != "_detalle"},
    "consecutivo": int(consecutivo),
    "frente": datos["frente"],
    "actividades": [a for a in actividades_editadas.split("\n") if a.strip()][:28],
    "observaciones": [o for o in observaciones_editadas.split("\n") if o.strip()][:8],
    "motivos_disponibilidad": datos["motivos_disponibilidad"],
    "fotos": fotos_final,
    "modo_foto": "llenar",
}
check_true("datos_final sin la clave interna _detalle", "_detalle" not in datos_final)
check("el contexto trae las 10 cajas medidas", len(ctx["cajas"]), 10)
check_true("contexto liviano (<100 KB)", len(__import__("pickle").dumps(ctx)) < 100_000,
           f'{len(__import__("pickle").dumps(ctx))/1024:.0f} KB')

salida2 = cr.generar_informe(plantilla_bytes, datos_final)
wb2 = openpyxl.load_workbook(io.BytesIO(salida2), data_only=True)
inf2 = wb2[cr.SH_INFORME]
check("actividad editada a mano", inf2.cell(row=665, column=7).value, "Línea agregada a mano.")
check("actividades recortadas tras la edición", inf2.cell(row=666, column=7).value, None)

# modo "llenar" debe recortar: la imagen sigue midiendo lo mismo que la caja
img_llenar = openpyxl.load_workbook(io.BytesIO(salida2))[cr.SH_FOTOS]._images[1]
caja = cr.medir_slot(wb0, 0)
check("modo llenar respeta el tamaño de caja",
      (img_llenar.width, img_llenar.height), caja)

nombre = (f"{cr.consecutivo_a_fecha(int(consecutivo), cr.leer_dia1(wb0)):%Y-%m-%d}"
          f"_8000008746_ODS03_Informe_Diario_{int(consecutivo)}.xlsx")
check("nombre del archivo de salida", nombre,
      "2026-08-23_8000008746_ODS03_Informe_Diario_140.xlsx")

print("\n" + "=" * 72)
if fallos:
    print(f"FALLARON {len(fallos)} comprobaciones:")
    for f in fallos:
        print(f"  - {f}")
    sys.exit(1)
print("PIPELINE COMPLETO OK")
