"""
Diagnostica un export de FastField contra el parser.

Dice, hoja por hoja y columna por columna, qué reconoció el parser y qué no.
Sirve para ajustar el mapeo sin adivinar: se corre con un submission de prueba
y muestra exactamente dónde no coinciden los nombres.

    python3 scripts/diagnostico_export.py <export_de_fastfield.xlsx>
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from utils import cenit_report as cr
from utils.fastfield import (
    _METADATOS, _clasificar, _es_picker_fotos, _filas, norm, parse_submission,
)

# Para cada dato del informe, todas las etiquetas que el parser acepta.
# La primera es la de la especificación; las demás son las que ya se vieron en
# formularios reales.
ETIQUETAS = {
    "jornada": [
        ("Frente", ["Frente"]),
        ("Hora de inicio", ["Hora de inicio", "Hora inicio"]),
        ("Hora final", ["Hora final", "Hora fin"]),
        ("Total de horas trabajadas", ["Total de horas trabajadas", "Total horas"]),
        ("¿Hubo algún evento de suspensión?",
         ["Hubo algun evento de suspension", "Evento"]),
        ("Hora en que inició el evento",
         ["Hora en que inicio el evento", "Hora inicio evento"]),
        ("Hora en que terminó el evento",
         ["Hora en que termino el evento", "Hora fin evento"]),
        ("¿Qué ocurrió?", ["Que ocurrio", "Descripcion del evento"]),
    ],
    "items": [
        ("Ítem de pago", ["Item de pago ejecutado", "Item de pago", "Items cenit", "Item"]),
        ("Cantidad", ["Cantidad #1", "Cantidad"]),
        ("Cantidad — dimensión 2", ["Cantidad — dimensión 2", "Cantidad dimension 2", "Cantidad #2"]),
        ("Cantidad — dimensión 3", ["Cantidad — dimensión 3", "Cantidad dimension 3", "Cantidad #3"]),
    ],
    "mano_obra": [
        ("Cargo", ["Cargo"]),
        ("Cantidad de personas", ["Cantidad", "Cant"]),
        ("Horas laboradas", ["Horas laboradas", "Horas"]),
        ("Horas disponible", ["Horas disponible"]),
    ],
    "equipos": [
        ("Tipo de equipo", ["Tipo de equipo"]),
        ("Cantidad", ["Cantidad", "Cant"]),
        ("Horas laboradas", ["Horas laboradas", "Horas"]),
        ("Horas disponible", ["Horas disponible"]),
        ("Horas fuera de servicio", ["Horas fuera de servicio"]),
    ],
    "actividades": [
        ("Actividad", ["Describa la actividad ejecutada", "Actividad"]),
    ],
    "observaciones": [
        ("Observación", ["Observacion", "Observaciones"]),
    ],
}
ROOT_ESPERADAS = [
    ("Fecha del informe", ["Fecha del informe", "Fecha"]),
    ("Frente o sitio de trabajo",
     ["Frente o sitio de trabajo", "Frente / Sitio", "Frente", "Locacion", "Sitio"]),
    ("Motivos de disponibilidad", ["Motivos de disponibilidad", "Motivos disponibilidad"]),
    ("Profesional líder PCC", ["Profesional lider PCC", "Profesional Lider"]),
]
FOTO_ESPERADAS = [("Foto", ["Photo", "Foto"]),
                  ("Descripción", ["Comment", "Descripcion de la foto", "Descripcion"])]


def reconoce(headers, alternativas) -> str | None:
    """Header real que el parser asociaría, probando todas las alternativas."""
    for alt in alternativas:
        clave = norm(alt)
        for h in headers:
            if h and (norm(h) == clave or norm(h).startswith(clave)):
                return h
    return None


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    ruta = Path(sys.argv[1])
    if not ruta.exists():
        sys.exit(f"No existe: {ruta}")

    raw = ruta.read_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    print(f"\nExport: {ruta.name}")
    print(f"Hojas:  {wb.sheetnames}\n")
    print("=" * 74)

    hallados = set()
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        filas = _filas(ws)

        if norm(nombre) == "root":
            clase = "Root (datos generales)"
        elif _es_picker_fotos(headers, nombre):
            clase = "Registro fotográfico"
        else:
            c = _clasificar(headers)
            clase = f"{c}" if c else "NO RECONOCIDA"
        hallados.add(clase)

        marca = "  " if clase != "NO RECONOCIDA" else "!!"
        print(f"\n{marca} [{nombre}] -> {clase}   ({len(filas)} fila(s))")

        esperadas = (ROOT_ESPERADAS if norm(nombre) == "root"
                     else FOTO_ESPERADAS if clase == "Registro fotográfico"
                     else ETIQUETAS.get(clase, []))

        usados = set()
        for nombre_dato, alternativas in esperadas:
            real = reconoce(headers, alternativas)
            if real:
                usados.add(real)
                muestra = ""
                if filas:
                    v = filas[0].get(real)
                    muestra = f"   ej: {str(v)[:36]!r}" if v not in (None, "") else "   (vacío)"
                print(f"     OK    {nombre_dato:34s} <- {str(real)[:30]!r}{muestra}")
            else:
                print(f"     falta {nombre_dato:34s} <- sin columna (opcional si no la creaste)")

        sobrantes = [h for h in headers
                     if h and h not in usados and norm(h) not in _METADATOS]
        if sobrantes:
            print(f"     SIN USAR: {[str(h)[:26] for h in sobrantes]}")

    print("\n" + "=" * 74)
    print("\nRESULTADO DEL PARSEO COMPLETO\n")
    cap = parse_submission(io.BytesIO(raw))
    for k in ("fecha", "frente", "motivos_disponibilidad"):
        print(f"  {k:24s} {cap[k]!r}")
    for k in ("actividades", "items", "mano_obra", "equipos", "jornadas",
              "observaciones", "fotos"):
        n = len(cap[k])
        marca = "  " if n else "!!"
        print(f"{marca} {k:24s} {n}")
    if cap["avisos"]:
        print("\n  Avisos:")
        for a in cap["avisos"]:
            print(f"    · {a}")

    # ── Resolución contra el libro maestro, si está a mano ──────────────────
    plantilla = Path.home() / "Downloads" / "2026-08-22_8000008746_ODS03_Informe_Diario_139.xlsx"
    if plantilla.exists() and cap["fecha"]:
        from utils import armado
        ctx = cr.construir_contexto(plantilla.read_bytes())
        consec = cr.fecha_a_consecutivo(cap["fecha"], ctx["dia1"])
        print(f"\nRESOLUCIÓN CONTRA EL LIBRO (consecutivo {consec})\n")
        datos, avisos = armado.construir(cap, ctx, max(1, consec))
        det = datos["_detalle"]
        print(f"  ítems resueltos     {len(det['items'])}/{len(cap['items'])}")
        print(f"  cargos resueltos    {len(det['mano_obra'])}/{len(cap['mano_obra'])}")
        print(f"  equipos resueltos   {len(det['equipos'])}/{len(cap['equipos'])}")
        if avisos:
            print("\n  Avisos:")
            for a in avisos:
                print(f"    · {a}")

    faltan = [c for c in ETIQUETAS if c not in hallados]
    if faltan:
        print(f"\nSECCIONES NO ENCONTRADAS EN EL EXPORT: {faltan}")
        print("Puede ser que esas páginas fueran vacías en este submission de prueba,")
        print("o que los nombres de sus campos no coincidan. Revisa el detalle arriba.")


if __name__ == "__main__":
    main()
