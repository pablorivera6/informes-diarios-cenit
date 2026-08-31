"""
Prueba de humo del motor de Excel contra el informe real de Cenit.

Genera el informe 140 (23-ago-2026) a partir del 139 y verifica celda por celda
que todo quedó donde debe. No necesita FastField ni credenciales.

    python3 scripts/smoke_test.py
"""
import io
import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.utils import get_column_letter as gl
from PIL import Image, ImageDraw

from utils import cenit_report as cr

PLANTILLA = Path.home() / "Downloads" / "2026-08-22_8000008746_ODS03_Informe_Diario_139.xlsx"
SALIDA    = Path(__file__).resolve().parent.parent / "salida_prueba_140.xlsx"

fallos: list[str] = []


def check(nombre: str, obtenido, esperado):
    ok = obtenido == esperado
    print(f"  {'PASA' if ok else 'FALLA'}  {nombre}: {obtenido!r}"
          + ("" if ok else f"  (esperado {esperado!r})"))
    if not ok:
        fallos.append(nombre)


def check_true(nombre: str, cond, detalle=""):
    print(f"  {'PASA' if cond else 'FALLA'}  {nombre}" + (f": {detalle}" if detalle else ""))
    if not cond:
        fallos.append(nombre)


def foto_falsa(texto: str, size=(1600, 1200)) -> bytes:
    """Genera un JPEG de prueba con proporción distinta a la caja, para
    comprobar que se encaja sin deformar."""
    img = Image.new("RGB", size, (38, 70, 83))
    d = ImageDraw.Draw(img)
    d.rectangle([40, 40, size[0] - 40, size[1] - 40], outline=(233, 196, 106), width=14)
    d.text((size[0] // 2 - 90, size[1] // 2), texto, fill=(255, 255, 255))
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=88)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════

print(f"\nPlantilla: {PLANTILLA.name}")
if not PLANTILLA.exists():
    sys.exit(f"No se encontró la plantilla en {PLANTILLA}")

plantilla_bytes = PLANTILLA.read_bytes()
wb0 = cr.abrir_libro(plantilla_bytes)

dia1        = cr.leer_dia1(wb0)
consec_prev = cr.leer_consecutivo_actual(wb0)
items       = cr.leer_items(wb0)
cargos      = cr.leer_cargos(wb0)
equipos     = cr.leer_equipos(wb0)

NUEVO = consec_prev + 1
fecha_nueva = cr.consecutivo_a_fecha(NUEVO, dia1)

print("\n── Lectura de la plantilla ──────────────────────────────────────────")
check("dia1 del proyecto", dia1.isoformat(), "2026-04-06")
check("consecutivo actual", consec_prev, 139)
check("fecha del consecutivo 139", cr.consecutivo_a_fecha(139, dia1).isoformat(), "2026-08-22")
check("ida y vuelta fecha->consecutivo", cr.fecha_a_consecutivo(fecha_nueva, dia1), NUEVO)
check_true("ítems leídos de Costo Real PDT", 85 <= len(items) <= 95,
           f"{len(items)} ítems (el catálogo real son ~92 filas, no las 597 de RESUMEN PXQ)")
check_true("cargos leídos de HH", len(cargos) >= 10, f"{len(cargos)} cargos")
check_true("equipos leídos de EQUIPOS", len(equipos) >= 11, f"{len(equipos)} equipos")

col_costo = cr.columna_del_dia(cr.SH_COSTO, NUEVO)
col_hh    = cr.columna_del_dia(cr.SH_HH, NUEVO)
col_eq    = cr.columna_del_dia(cr.SH_EQUIPOS, NUEVO)
print(f"\n  Columnas del día {NUEVO} ({fecha_nueva}): "
      f"Costo={gl(col_costo)}  HH={gl(col_hh)}  EQUIPOS={gl(col_eq)}")

# La columna calculada debe coincidir con la fecha que trae el encabezado
check("fecha en encabezado de Costo Real PDT",
      wb0[cr.SH_COSTO].cell(row=3, column=col_costo).value.date().isoformat(),
      fecha_nueva.isoformat())
check("fecha en encabezado de HH",
      wb0[cr.SH_HH].cell(row=5, column=col_hh).value.date().isoformat(),
      fecha_nueva.isoformat())
check("fecha en encabezado de EQUIPOS",
      wb0[cr.SH_EQUIPOS].cell(row=5, column=col_eq).value.date().isoformat(),
      fecha_nueva.isoformat())

# ══ Datos de prueba ═══════════════════════════════════════════════════════════

item_a = next(i for i in items if not i["es_encabezado"] and i["unidad"] == "m")
item_b = next(i for i in items if not i["es_encabezado"] and i["num"] != item_a["num"])

cargo_res = next(c for c in cargos if "Residente" in c["cargo"])
cargo_ing = next(c for c in cargos if c["cargo"] == "Ingeniero")
eq_camio  = next(e for e in equipos if "Camioneta" in e["tipo"])

ACTIVIDADES = [
    "Instalación y tendido de cable de protección catódica en PK-152+428.",
    "Excavación manual para caja de inspección, frente 1.",
    "Verificación de continuidad eléctrica en estación de prueba.",
]
OBSERVACIONES = [
    "Se presentó lluvia entre las 14:00 y 15:00, se suspendió actividad en zanja.",
    "Interventoría acompañó la jornada en campo.",
]

datos = {
    "consecutivo": NUEVO,
    "frente": "PK-152+428",
    "actividades": ACTIVIDADES,
    "observaciones": OBSERVACIONES,
    "motivos_disponibilidad": "Retroexcavadora en mantenimiento preventivo programado.",
    "avances": [
        {"row_num": item_a["row_num"], "cantidad": 12.5},
        {"row_num": item_b["row_num"], "cantidad": 0.25},
    ],
    "hh": [
        {"row_num": cargo_res["row_num"], "horas": 9},
        {"row_num": cargo_ing["row_num"], "horas": 8},
    ],
    "equipos": [
        {"row_num": eq_camio["row_num"], "horas": 10},
    ],
    "jornadas": [
        {"frente": "PK-152+428", "hora_inicio": "07:00", "hora_final": "17:00",
         "total_horas": "9:00", "hubo_evento": "SI", "evento_inicio": "14:00",
         "evento_fin": "15:00", "evento_desc": "Suspensión por tormenta eléctrica."},
        {"frente": "PK-47+128", "hora_inicio": "07:30", "hora_final": "16:30",
         "total_horas": "9:00", "hubo_evento": "NO"},
    ],
    "eq_fuera_servicio": [{"row_num": 696, "horas": 8}],
    "fotos": [
        {"image_bytes": foto_falsa("FOTO 1"), "descripcion": "Tendido de cable, vista general."},
        {"image_bytes": foto_falsa("FOTO 2"), "descripcion": "Detalle de la conexión termosoldada."},
        {"image_bytes": foto_falsa("FOTO 3"), "descripcion": "Excavación manual frente 1."},
    ],
}

print("\n── Generando informe ────────────────────────────────────────────────")
salida_bytes = cr.generar_informe(plantilla_bytes, datos)
SALIDA.write_bytes(salida_bytes)
print(f"  Generado: {SALIDA.name}  ({len(salida_bytes)/1024/1024:.2f} MB, "
      f"origen {len(plantilla_bytes)/1024/1024:.2f} MB)")

# ══ Verificación ══════════════════════════════════════════════════════════════

print("\n── Integridad del paquete ───────────────────────────────────────────")
zin  = zipfile.ZipFile(io.BytesIO(plantilla_bytes))
zout = zipfile.ZipFile(io.BytesIO(salida_bytes))
check_true("ZIP válido", zout.testzip() is None)
faltantes = set(zin.namelist()) - set(zout.namelist())
check("ninguna parte original perdida", faltantes, set())
nuevas = set(zout.namelist()) - set(zin.namelist())
check("partes nuevas de media", sorted(nuevas),
      ["xl/media/image4.png", "xl/media/image5.png", "xl/media/image6.png"])

wbx = zout.read("xl/workbook.xml").decode("utf-8")
check_true("fullCalcOnLoad activado", 'fullCalcOnLoad="1"' in wbx,
           re.search(r"<calcPr[^>]*/>", wbx).group(0))

drw = zout.read("xl/drawings/drawing1.xml").decode("utf-8")
check("anclajes de foto insertados", drw.count(cr.FOTO_NAME_PREFIX), 3)
check("logo original conservado", drw.count('name="1 Imagen"'), 1)
rels = zout.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
check("relaciones en el drawing", rels.count("<Relationship "), 4)

print("\n── Contenido del informe ────────────────────────────────────────────")
wb = openpyxl.load_workbook(io.BytesIO(salida_bytes), data_only=True)
inf, cos, hh, eq, fot = (wb[cr.SH_INFORME], wb[cr.SH_COSTO], wb[cr.SH_HH],
                         wb[cr.SH_EQUIPOS], wb[cr.SH_FOTOS])

check("Z6 consecutivo", inf.cell(row=6, column=26).value, NUEVO)
check("avance ítem A", cos.cell(row=item_a["row_num"], column=col_costo).value, 12.5)
check("avance ítem B", cos.cell(row=item_b["row_num"], column=col_costo).value, 0.25)
check("HH residente", hh.cell(row=cargo_res["row_num"], column=col_hh).value, 9)
check("HH ingeniero", hh.cell(row=cargo_ing["row_num"], column=col_hh).value, 8)
check("horas camioneta", eq.cell(row=eq_camio["row_num"], column=col_eq).value, 10)

check("frente (B662)", inf.cell(row=662, column=2).value, "PK-152+428")
for i, texto in enumerate(ACTIVIDADES):
    check(f"actividad {i+1} (G{662+i})", inf.cell(row=662 + i, column=7).value, texto)
# La plantilla (informe 139) traía texto en G663-G666 y G689: debe quedar limpio
for fila in (665, 666, 689):
    check(f"actividad de ayer borrada (G{fila})", inf.cell(row=fila, column=7).value, None)

check("jornada 1 frente", inf.cell(row=718, column=2).value, "PK-152+428")
check("jornada 1 hora inicio", inf.cell(row=718, column=7).value, "07:00")
check("jornada 1 evento desc", inf.cell(row=718, column=19).value,
      "Suspensión por tormenta eléctrica.")
check("jornada 2 frente", inf.cell(row=719, column=2).value, "PK-47+128")
check("jornada 2 sin evento", inf.cell(row=719, column=13).value, "NO")

check("motivos disponibilidad", inf.cell(row=713, column=2).value,
      "Retroexcavadora en mantenimiento preventivo programado.")
check("equipo fuera de servicio (AB696)", inf.cell(row=696, column=28).value, 8)
for i, texto in enumerate(OBSERVACIONES):
    check(f"observación {i+1} (B{722+i})", inf.cell(row=722 + i, column=2).value, texto)

check("desc foto 1 (F23)", fot.cell(row=23, column=6).value, "Tendido de cable, vista general.")
check("desc foto 2 (U23)", fot.cell(row=23, column=21).value,
      "Detalle de la conexión termosoldada.")
check("desc foto 3 (F36)", fot.cell(row=36, column=6).value, "Excavación manual frente 1.")

print("\n── Fórmulas preservadas ─────────────────────────────────────────────")
wbf = openpyxl.load_workbook(io.BytesIO(salida_bytes), data_only=False)
inff, cosf = wbf[cr.SH_INFORME], wbf[cr.SH_COSTO]
check("Z7 sigue siendo HLOOKUP",
      str(inff.cell(row=7, column=26).value).startswith("=+HLOOKUP"), True)
check("F14 sigue apuntando a RESUMEN PXQ",
      inff.cell(row=14, column=6).value, "='RESUMEN PXQ'!G11")
check("M98 de Costo Real PDT intacta",
      cosf.cell(row=98, column=13).value, "=SUM(N98:IIU98)")

print("\n── Guardia de reproceso ─────────────────────────────────────────────")
ya = cr.dia_ya_tiene_datos(wb, NUEVO)
check("detecta avances ya cargados en el día", ya["avances"], 2)
check("detecta HH ya cargadas en el día", ya["hh"], 2)
check("día limpio en la plantilla original", cr.dia_ya_tiene_datos(wb0, NUEVO)["avances"], 0)

print("\n── Avisos de datos manuales ─────────────────────────────────────────")
avisos = cr.detectar_datos_manuales(wb0)
check_true("detecta las filas 710/711 escritas a mano", len(avisos) >= 2, f"{len(avisos)} avisos")
for a in avisos:
    print(f"     · {a}")

print("\n── Idempotencia (regenerar sobre la salida) ─────────────────────────")
segunda = cr.generar_informe(salida_bytes, datos)
drw2 = zipfile.ZipFile(io.BytesIO(segunda)).read("xl/drawings/drawing1.xml").decode("utf-8")
check("no se duplican anclajes al regenerar", drw2.count(cr.FOTO_NAME_PREFIX), 3)
wb2 = openpyxl.load_workbook(io.BytesIO(segunda), data_only=True)
check("avance se ACUMULA al reprocesar el mismo día (esperado; la app avisa)",
      wb2[cr.SH_COSTO].cell(row=item_a["row_num"], column=col_costo).value, 25.0)
check("actividades no se duplican al regenerar",
      wb2[cr.SH_INFORME].cell(row=662 + len(ACTIVIDADES), column=7).value, None)

print("\n" + "=" * 72)
if fallos:
    print(f"FALLARON {len(fallos)} comprobaciones:")
    for f in fallos:
        print(f"  - {f}")
    sys.exit(1)
print("TODAS LAS COMPROBACIONES PASARON")
